
"""
Script:
    02_append_dippr_sponsor_sheet.py

Purpose:
    This script adds the DIPPR Sponsor component list Excel data
    as a new sheet into the existing consolidated library Excel workbook.

Functionality:
    - Reads DIPPR Sponsor element/component data from Excel
    - Opens the existing library Excel workbook
    - Checks for existing sheet names to avoid duplication
    - Creates a unique sheet name if required
    - Appends the DIPPR Sponsor data as a new sheet

Input:
    - DIPPR Sponsor Excel file
    - Existing library Excel workbook

Output:
    Updated Excel workbook containing the newly added DIPPR Sponsor sheet
"""

from pathlib import Path
from shutil import copyfile

import pandas as pd
from openpyxl import load_workbook

from pathlib import Path

# ==================================================
# CONFIGURATION
# ==================================================
RUN_YEAR = "2025"

ROOT_DIR = Path(r"D:\NIST_XML_Converter")

# ==================================================
# PREREQUISITE DIRECTORIES
# ==================================================
PREREQ_DIR = ROOT_DIR / "prerequisites"

EXCEL_INPUT_DIR = PREREQ_DIR / "excel_inputs"

# ==================================================
# OUTPUT DIRECTORIES
# ==================================================
OUTPUT_DIR = ROOT_DIR / "output" / RUN_YEAR

PROCESSED_DIR = OUTPUT_DIR / "processed" / "full_library"

XML_OUTPUT_DIR = OUTPUT_DIR / "xml"

# ==================================================
# INPUT / OUTPUT FILES
# ==================================================
elements_file = (
    EXCEL_INPUT_DIR
    / "1_DIPPR_SponsorListofElements.xlsx"
)

library_file = (
    PROCESSED_DIR
    / "1_Libraries_XML_Component_Extract_nodipprsponsor.xlsx"
)

output_library_file = (
    PROCESSED_DIR
    / "2_Libraries_XML_Component_Extract.xlsx"
)

# ==================================================
# SHEET CONFIGURATION
# ==================================================
base_sheet_name = "Dippr_Sponsor"

# ==================================================
# LOAD ELEMENT DATA
# ==================================================
elements_df = pd.read_excel(elements_file)

# ==================================================
# COPY ORIGINAL WORKBOOK TO NEW OUTPUT FILE
# ==================================================
copyfile(library_file, output_library_file)

# ==================================================
# LOAD WORKBOOK TO CHECK EXISTING SHEETS
# ==================================================
workbook = load_workbook(output_library_file)

sheet_name = base_sheet_name
i = 1

while sheet_name in workbook.sheetnames:
    sheet_name = f"{base_sheet_name}_{i}"
    i += 1

# ==================================================
# APPEND NEW SHEET
# ==================================================
with pd.ExcelWriter(
    output_library_file,
    engine="openpyxl",
    mode="a",
    if_sheet_exists="new"
) as writer:

    elements_df.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False
    )

print(f"Added new sheet '{sheet_name}' to:")
print(output_library_file)
